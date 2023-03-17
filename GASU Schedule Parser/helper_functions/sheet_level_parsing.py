from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import cell as xlcell
from consts.days_of_week import weekdays
from consts.lecture_time_ranges import lecture_time_ranges, first_time_range, second_time_range, third_time_range
from consts.lecture_time_ranges import fourth_time_range, fifth_time_range, sixth_time_range, seventh_time_range
from db_api.model.schedule_group_type_day import ScheduleGroupTypeDay



async def parse_worksheet(worksheet: Worksheet, academic_degree, faculty, year, dryrun=True):
    start_column_index, end_column_index, groupname_row_index = get_column_indexes_and_groupname_row_index(worksheet)
    time_range_width, start_row_index, time_ranges_in_a_day = \
        get_time_range_width_and_start_row_index_and_timeranges_count(worksheet)

    for column_index in range(start_column_index, end_column_index + 1):
        await parse_schedule_for_one_group(worksheet=worksheet, column_index=column_index,
                                     groupname_row_index=groupname_row_index, start_row_index=start_row_index,
                                     time_range_width=time_range_width, time_ranges_in_a_day=time_ranges_in_a_day,
                                     academic_degree=academic_degree, faculty=faculty, year=year,
                                     dryrun=dryrun)




async def parse_schedule_for_one_group(worksheet: Worksheet, column_index, groupname_row_index, start_row_index,
                                       time_range_width, time_ranges_in_a_day,
                                       academic_degree, faculty, year, dryrun=True):

    accidental_service_words_or_empty = ('', ' ', '_', '-', None)

    #Расписание на дни недели для недели "ЧИСЛИТЕЛЬ"
    group_even_week_schedules = []
    EVEN_WEEK_TYPE = 'ЧИСЛИТЕЛЬ'

    # Расписание на дни недели для недели "ЗНАМЕНАТЕЛЬ"
    group_odd_week_schedules = []
    ODD_WEEK_TYPE = 'ЗНАМЕНАТЕЛЬ'

    group_name = str(worksheet.cell(row=groupname_row_index, column=column_index).value)
    if group_name == "None":
        return
    # Sometimes group_name is written in a cell like 'ГР.3-А-2', sometimes like 'ГРУППА 3-А-2'
    if len(group_name.split(" ")) > 1:
        group_name = group_name.split(" ")[-1]
    else:
        group_name = group_name.split(".")[-1]

    group_schedules = []

    row_index = start_row_index

    end_row_index = worksheet.max_row

    # day_info[0] is name, [1] - index 1-6
    for day_info in weekdays:

        time_ranges_in_a_day = get_timeranges_count(worksheet, day_info)
        id_even = group_name + "_0_" + str(day_info[1])
        id_odd = group_name + "_1_" + str(day_info[1])

        schedule_group_day_even_week = ScheduleGroupTypeDay(
            group_name_with_type_and_day=id_even,
            academic_degree=academic_degree,
            faculty=faculty,
            year=year,
            group_name=group_name,
            week_type=EVEN_WEEK_TYPE,
            day_of_week=day_info[0],
            first_lecture_time=first_time_range,
            second_lecture_time=second_time_range,
            third_lecture_time=third_time_range,
            fourth_lecture_time=fourth_time_range,
            fifth_lecture_time=fifth_time_range
        )

        schedule_group_day_odd_week = ScheduleGroupTypeDay(
            group_name_with_type_and_day=id_odd,
            academic_degree=academic_degree,
            faculty=faculty,
            year=year,
            group_name=group_name,
            week_type=ODD_WEEK_TYPE,
            day_of_week=day_info[0],
            first_lecture_time=first_time_range,
            second_lecture_time=second_time_range,
            third_lecture_time=third_time_range,
            fourth_lecture_time=fourth_time_range,
            fifth_lecture_time=fifth_time_range
        )

        if time_ranges_in_a_day > 5:
            schedule_group_day_even_week.sixth_lecture_time = sixth_time_range
            schedule_group_day_odd_week.sixth_lecture_time = sixth_time_range
            schedule_group_day_even_week.seventh_lecture_time = seventh_time_range
            schedule_group_day_odd_week.seventh_lecture_time = seventh_time_range

        time_range_counter = 1

        time_ranges = lecture_time_ranges
        if time_ranges_in_a_day == 5:
            time_ranges = time_ranges[:-2]
        if time_ranges_in_a_day == 6:
            time_ranges = time_ranges[:-1]

        for time_range in time_ranges:
            even_lectures = []
            odd_lectures = []
            # 1 time range block length in rows
            for i in range(time_range_width):
                cell_value = get_cell_value(worksheet,
                                            worksheet.cell(row=row_index, column=column_index))
                if cell_value not in accidental_service_words_or_empty:
                    if isinstance(cell_value, str):
                        cell_value = remove_nextline_symbol(cell_value)
                    if i < (time_range_width // 2):
                        if str(cell_value).strip() not in even_lectures:
                            even_lectures.append(str(cell_value).strip())
                    else:
                        if str(cell_value).strip() not in odd_lectures:
                            odd_lectures.append(str(cell_value).strip())
                row_index += 1

            # getting lecture(s)
            if len(even_lectures) > 0:
                even_lecture_full_name = " ".join(even_lectures)
            else:
                even_lecture_full_name = "Нет занятий."

            if len(odd_lectures) > 0:
                odd_lecture_full_name = " ".join(odd_lectures)
            else:
                odd_lecture_full_name = "Нет занятий."

            if time_range_counter == 1:
                schedule_group_day_even_week.first_lecture = even_lecture_full_name
                schedule_group_day_odd_week.first_lecture = odd_lecture_full_name
            elif time_range_counter == 2:
                schedule_group_day_even_week.second_lecture = even_lecture_full_name
                schedule_group_day_odd_week.second_lecture = odd_lecture_full_name
            elif time_range_counter == 3:
                schedule_group_day_even_week.third_lecture = even_lecture_full_name
                schedule_group_day_odd_week.third_lecture = odd_lecture_full_name
            elif time_range_counter == 4:
                schedule_group_day_even_week.fourth_lecture = even_lecture_full_name
                schedule_group_day_odd_week.fourth_lecture = odd_lecture_full_name
            elif time_range_counter == 5:
                schedule_group_day_even_week.fifth_lecture = even_lecture_full_name
                schedule_group_day_odd_week.fifth_lecture = odd_lecture_full_name
            elif time_range_counter == 6:
                schedule_group_day_even_week.sixth_lecture = even_lecture_full_name
                schedule_group_day_odd_week.sixth_lecture = odd_lecture_full_name
            elif time_range_counter == 7:
                schedule_group_day_even_week.seventh_lecture = even_lecture_full_name
                schedule_group_day_odd_week.seventh_lecture = odd_lecture_full_name

            if time_range_counter == time_ranges_in_a_day:
                # if we're in last time range, skip 1 empty row before next day, append schedule_group_day to list
                row_index += 1
                group_schedules.append(schedule_group_day_even_week)
                group_schedules.append(schedule_group_day_odd_week)
                if dryrun:
                    print(schedule_group_day_even_week)
                    print(schedule_group_day_odd_week)
                else:
                    await schedule_group_day_even_week.create()
                    await schedule_group_day_odd_week.create()


            # append to list on last row
            # on saturday there are less lectures
            if row_index == end_row_index:
                group_schedules.append(schedule_group_day_even_week)
                group_schedules.append(schedule_group_day_odd_week)
                break

            time_range_counter += 1
            even_lectures.clear()
            odd_lectures.clear()












# section get cell value
def within_range(bounds: tuple, cell: xlcell) -> bool:
    column_start, row_start, column_end, row_end = bounds
    row = cell.row
    if row >= row_start and row <= row_end:
        column = cell.column
        if column >= column_start and column <= column_end:
            return True
    return False


def get_cell_value(sheet: Worksheet, cell: xlcell) -> any:
    for merged in sheet.merged_cells:
        if within_range(merged.bounds, cell):
            return sheet.cell(merged.min_row, merged.min_col).value
    return cell.value



# section indexes
def get_column_indexes_and_groupname_row_index(target_worksheet: Worksheet):
    for row in target_worksheet.iter_rows():
        for cell in row:
            if str(cell.value).lower().startswith("время занятий"):
                groupname_row_index = cell.row
                starting_column_index = cell.column + 2
                ending_column_index = target_worksheet.max_column
                return starting_column_index, ending_column_index, groupname_row_index



def get_time_range_width_and_start_row_index_and_timeranges_count(target_worksheet: Worksheet):
    day_start_row = 0
    day_end_row = 0
    start_row_index = 0
    MONDAY_NAMES = ("ПОНЕДЕЛЬНИК", "понедельник", "Понедельник")
    FIFTH_TIME_RANGE = '1645-1815'
    for row in target_worksheet.iter_rows(max_col=1):
        for cell in row:
            cell_value = get_cell_value(target_worksheet, cell)
            if (day_start_row == 0) and (cell_value in MONDAY_NAMES):
                day_start_row = cell.row
                start_row_index = cell.row
                continue
            if (day_start_row != 0) and (cell_value not in MONDAY_NAMES):
                day_end_row = cell.row
                day_width = day_end_row - day_start_row
                time_range = get_cell_value(target_worksheet,
                                            target_worksheet.cell(row=(cell.row - 1), column=(cell.column + 2)))
                if str(time_range) == FIFTH_TIME_RANGE:
                    time_ranges_in_a_day = 5
                else:
                    time_ranges_in_a_day = 7
                time_range_width = day_width // time_ranges_in_a_day
                return time_range_width, start_row_index, time_ranges_in_a_day


def get_timeranges_count(target_worksheet: Worksheet, day_info):
    day_start_row = 0
    day_end_row = 0
    start_row_index = 0
    DAY_NAMES = day_info[0].upper()
    FIFTH_TIME_RANGE = '1645-1815'
    SIXTH_TIME_RANGE = '1830-2000'
    time_ranges_in_a_day = 5
    if DAY_NAMES == "СУББОТА":
        return 5
    for row in target_worksheet.iter_rows(max_col=1):
        for cell in row:
                cell_value = get_cell_value(target_worksheet, cell)
                if (day_start_row == 0) and (cell_value == DAY_NAMES):
                    day_start_row = cell.row
                    start_row_index = cell.row
                    continue
                if (day_start_row != 0) and (cell_value != DAY_NAMES):
                    day_end_row = cell.row
                    day_width = day_end_row - day_start_row
                    time_range = get_cell_value(target_worksheet,
                                            target_worksheet.cell(row=(cell.row - 1), column=(cell.column + 2)))
                    
                    if str(time_range) == FIFTH_TIME_RANGE:
                        time_ranges_in_a_day = 5
                    elif str(time_range) == SIXTH_TIME_RANGE:
                        time_ranges_in_a_day = 6
                    else:
                        time_ranges_in_a_day = 7
                    time_range_width = day_width // time_ranges_in_a_day

                    return time_ranges_in_a_day


# section utils - fuction to remove nextline symbols
def remove_nextline_symbol(target_string):
    target_string = target_string.replace("\n", " ")
    return target_string




